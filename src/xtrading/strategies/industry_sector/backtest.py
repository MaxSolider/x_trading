"""
回测验证模块
基于指定日期的真实数据进行回测验证不同策略的表现
"""

import pandas as pd
import numpy as np
import os
from typing import Optional, Dict, Any, List, Tuple
from datetime import datetime, timedelta
from .macd_strategy import IndustryMACDStrategy
from .rsi_strategy import IndustryRSIStrategy
from .bollinger_bands_strategy import IndustryBollingerBandsStrategy
from .moving_average_strategy import IndustryMovingAverageStrategy
from ...repositories.stock.industry_info_query import IndustryInfoQuery
from ...static import INDUSTRY_CATEGORIES

class StrategyBacktest:
    """策略回测验证类"""
    
    def __init__(self):
        """初始化回测类"""
        self.industry_query = IndustryInfoQuery()
        self.macd_strategy = IndustryMACDStrategy()
        self.rsi_strategy = IndustryRSIStrategy()
        self.bb_strategy = IndustryBollingerBandsStrategy()
        self.ma_strategy = IndustryMovingAverageStrategy()
    
    def get_industry_category(self, industry_name: str) -> str:
        """
        根据行业名称获取对应的分类
        
        Args:
            industry_name: 行业板块名称
            
        Returns:
            str: 行业分类名称
        """
        for category, sectors in INDUSTRY_CATEGORIES.items():
            if industry_name in sectors:
                return category
        return "其他"  # 如果找不到分类，默认为"其他"
    
    def get_historical_data(self, industry_name: str, start_date: str, end_date: str) -> Optional[pd.DataFrame]:
        """
        获取行业板块历史数据
        
        Args:
            industry_name: 行业板块名称
            start_date: 开始日期 (YYYYMMDD)
            end_date: 结束日期 (YYYYMMDD)
            
        Returns:
            DataFrame: 历史数据
        """
        try:
            
            hist_data = self.industry_query.get_board_industry_hist(industry_name, start_date, end_date)
            if hist_data is None or hist_data.empty:
                print(f"❌ 无法获取 {industry_name} 的历史数据")
                return None
            
            print(f"✅ 成功获取 {industry_name} 历史数据 ({len(hist_data)} 条记录)")
            return hist_data
            
        except Exception as e:
            print(f"❌ 获取 {industry_name} 历史数据失败: {e}")
            return None
    
    def calculate_strategy_signals(self, data: pd.DataFrame, strategy_name: str, **kwargs) -> Optional[pd.DataFrame]:
        """
        计算策略交易信号
        
        Args:
            data: 历史数据
            strategy_name: 策略名称
            **kwargs: 策略参数
            
        Returns:
            DataFrame: 包含交易信号的DataFrame
        """
        try:
            if strategy_name == "MACD":
                # 提取MACD相关参数
                macd_params = {k: v for k, v in kwargs.items() if k in ['fast_period', 'slow_period', 'signal_period']}
                # 计算MACD指标
                macd_data = self.macd_strategy.calculate_macd(data, **macd_params)
                if macd_data is None:
                    return None
                # 生成交易信号
                return self.macd_strategy.generate_macd_signals(macd_data)
                
            elif strategy_name == "RSI":
                # 提取RSI相关参数
                rsi_params = {k: v for k, v in kwargs.items() if k in ['period']}
                signal_params = {k: v for k, v in kwargs.items() if k in ['oversold', 'overbought']}
                # 计算RSI指标
                rsi_data = self.rsi_strategy.calculate_rsi(data, **rsi_params)
                if rsi_data is None:
                    return None
                # 生成交易信号
                return self.rsi_strategy.generate_rsi_signals(rsi_data, **signal_params)
                
            elif strategy_name == "BollingerBands":
                # 提取布林带相关参数
                bb_params = {k: v for k, v in kwargs.items() if k in ['period', 'std_dev']}
                # 计算布林带指标
                bb_data = self.bb_strategy.calculate_bollinger_bands(data, **bb_params)
                if bb_data is None:
                    return None
                # 生成交易信号
                return self.bb_strategy.generate_bollinger_signals(bb_data)
                
            elif strategy_name == "MovingAverage":
                # 提取移动平均相关参数
                ma_params = {k: v for k, v in kwargs.items() if k in ['short_period', 'medium_period', 'long_period']}
                # 计算移动平均指标
                ma_data = self.ma_strategy.calculate_moving_averages(data, **ma_params)
                if ma_data is None:
                    return None
                # 生成交易信号
                return self.ma_strategy.generate_ma_signals(ma_data)
                
            else:
                print(f"❌ 不支持的策略: {strategy_name}")
                return None
                
        except Exception as e:
            print(f"❌ 计算 {strategy_name} 策略信号失败: {e}")
            return None
    
    def simulate_trading(self, signal_data: pd.DataFrame, initial_capital: float = 100000) -> Dict[str, Any]:
        """
        模拟交易过程
        
        Args:
            signal_data: 包含交易信号的DataFrame
            initial_capital: 初始资金
            
        Returns:
            Dict: 交易结果统计
        """
        if signal_data is None or signal_data.empty:
            return None
        
        # 确保有收盘价列
        close_col = None
        for col in ['收盘价', 'close', 'Close']:
            if col in signal_data.columns:
                close_col = col
                break
        
        if close_col is None:
            print("❌ 未找到收盘价列")
            return None
        
        # 初始化交易状态
        capital = initial_capital
        position = 0  # 持仓数量
        trades = []  # 交易记录
        portfolio_values = []  # 组合价值记录
        
        # 遍历每个交易日
        for i, row in signal_data.iterrows():
            current_price = row[close_col]
            signal = row.get('Signal', 0)
            signal_type = row.get('Signal_Type', 'HOLD')
            date = row.get('日期', f'Day_{i}')
            
            # 计算当前组合价值
            current_value = capital + (position * current_price)
            portfolio_values.append({
                'date': date,
                'price': current_price,
                'capital': capital,
                'position': position,
                'portfolio_value': current_value,
                'signal': signal,
                'signal_type': signal_type
            })
            
            # 执行交易
            if signal == 1:  # 买入信号
                if capital > 0:
                    # 全仓买入
                    shares_to_buy = capital // current_price
                    if shares_to_buy > 0:
                        cost = shares_to_buy * current_price
                        capital -= cost
                        position += shares_to_buy
                        trades.append({
                            'date': date,
                            'action': 'BUY',
                            'price': current_price,
                            'shares': shares_to_buy,
                            'amount': cost,
                            'capital_after': capital,
                            'position_after': position
                        })
            
            elif signal == -1:  # 卖出信号
                if position > 0:
                    # 全仓卖出
                    proceeds = position * current_price
                    capital += proceeds
                    trades.append({
                        'date': date,
                        'action': 'SELL',
                        'price': current_price,
                        'shares': position,
                        'amount': proceeds,
                        'capital_after': capital,
                        'position_after': 0
                    })
                    position = 0
            
            elif signal == 2:  # 强势买入
                if capital > 0:
                    shares_to_buy = capital // current_price
                    if shares_to_buy > 0:
                        cost = shares_to_buy * current_price
                        capital -= cost
                        position += shares_to_buy
                        trades.append({
                            'date': date,
                            'action': 'STRONG_BUY',
                            'price': current_price,
                            'shares': shares_to_buy,
                            'amount': cost,
                            'capital_after': capital,
                            'position_after': position
                        })
            
            elif signal == -2:  # 强势卖出
                if position > 0:
                    proceeds = position * current_price
                    capital += proceeds
                    trades.append({
                        'date': date,
                        'action': 'STRONG_SELL',
                        'price': current_price,
                        'shares': position,
                        'amount': proceeds,
                        'capital_after': capital,
                        'position_after': 0
                    })
                    position = 0
        
        # 计算最终结果
        final_price = signal_data.iloc[-1][close_col]
        final_value = capital + (position * final_price)
        
        # 计算统计指标
        portfolio_df = pd.DataFrame(portfolio_values)
        returns = portfolio_df['portfolio_value'].pct_change().dropna()
        
        total_return = (final_value - initial_capital) / initial_capital
        annualized_return = (1 + total_return) ** (252 / len(signal_data)) - 1
        volatility = returns.std() * np.sqrt(252)
        sharpe_ratio = annualized_return / volatility if volatility > 0 else 0
        max_drawdown = self.calculate_max_drawdown(portfolio_df['portfolio_value'])
        
        return {
            'initial_capital': initial_capital,
            'final_value': final_value,
            'total_return': total_return,
            'annualized_return': annualized_return,
            'volatility': volatility,
            'sharpe_ratio': sharpe_ratio,
            'max_drawdown': max_drawdown,
            'total_trades': len(trades),
            'buy_trades': len([t for t in trades if t['action'] in ['BUY', 'STRONG_BUY']]),
            'sell_trades': len([t for t in trades if t['action'] in ['SELL', 'STRONG_SELL']]),
            'trades': trades,
            'portfolio_values': portfolio_values
        }
    
    def calculate_max_drawdown(self, portfolio_values: pd.Series) -> float:
        """
        计算最大回撤
        
        Args:
            portfolio_values: 组合价值序列
            
        Returns:
            float: 最大回撤
        """
        peak = portfolio_values.expanding().max()
        drawdown = (portfolio_values - peak) / peak
        return drawdown.min()
    
    def backtest_strategy(self, industry_name: str, strategy_name: str, 
                         start_date: str, end_date: str, initial_capital: float = 100000,
                         **strategy_params) -> Optional[Dict[str, Any]]:
        """
        回测单个策略
        
        Args:
            industry_name: 行业板块名称
            strategy_name: 策略名称
            start_date: 开始日期
            end_date: 结束日期
            initial_capital: 初始资金
            **strategy_params: 策略参数
            
        Returns:
            Dict: 回测结果
        """
        print(f"🔍 开始回测 {industry_name} 的 {strategy_name} 策略...")
        
        # 获取历史数据
        hist_data = self.get_historical_data(industry_name, start_date, end_date)
        if hist_data is None:
            return None
        
        # 计算策略信号
        signal_data = self.calculate_strategy_signals(hist_data, strategy_name, **strategy_params)
        if signal_data is None:
            return None
        
        # 模拟交易
        trading_result = self.simulate_trading(signal_data, initial_capital)
        if trading_result is None:
            return None
        
        # 添加策略信息
        trading_result.update({
            'industry_name': industry_name,
            'strategy_name': strategy_name,
            'start_date': start_date,
            'end_date': end_date,
            'strategy_params': strategy_params,
            'data_points': len(hist_data)
        })
        
        print(f"✅ {industry_name} {strategy_name} 策略回测完成")
        return trading_result
    
    def compare_strategies(self, industry_name: str, strategies: List[str], 
                         start_date: str, end_date: str, initial_capital: float = 100000,
                         **strategy_params) -> List[Dict[str, Any]]:
        """
        比较多个策略的表现
        
        Args:
            industry_name: 行业板块名称
            strategies: 策略名称列表
            start_date: 开始日期
            end_date: 结束日期
            initial_capital: 初始资金
            **strategy_params: 策略参数
            
        Returns:
            List[Dict]: 各策略回测结果列表
        """
        results = []
        
        print(f"🔍 开始比较 {industry_name} 的 {len(strategies)} 个策略...")
        
        for strategy in strategies:
            result = self.backtest_strategy(industry_name, strategy, start_date, end_date, 
                                          initial_capital, **strategy_params)
            if result:
                results.append(result)
        
        # 按总收益率排序
        results.sort(key=lambda x: x['total_return'], reverse=True)
        
        print(f"✅ {industry_name} 策略比较完成")
        return results
    
    def print_backtest_results(self, results: List[Dict[str, Any]]):
        """
        打印回测结果（使用DataFrame格式）
        
        Args:
            results: 回测结果列表
        """
        if not results:
            print("❌ 无回测结果可显示")
            return
        
        print("\n" + "=" * 80)
        print("策略回测结果")
        print("=" * 80)
        print(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"行业板块: 行业板块分析")
        print(f"回测期间: {results[0]['start_date']} 至 {results[0]['end_date']}")
        print(f"策略数量: {len(results)}")
        print()
        
        # 1. 综合结果表
        self._print_comprehensive_table(results)
        
        # 2. 日收益明细表
        self._print_daily_returns_details(results)
        
        # 3. 累计收益明细表
        self._print_cumulative_returns_details(results)
        
        # 创建按行业分类的回测结果目录
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        industry_name = results[0].get('industry_name', 'Unknown')
        category = self.get_industry_category(industry_name)
        
        # 创建分类目录结构: tests/backtest/{行业分类名称}/{行业板块名称}/{时间戳}
        result_dir = f"tests/backtest/{category}/{industry_name}/{timestamp}"
        os.makedirs(result_dir, exist_ok=True)
        
        print(f"📁 回测结果目录: {result_dir}")
        
        # 3. 保存Markdown报告
        self._save_markdown_report(results, result_dir)
        
        # 4. 生成Excel文件
        self._generate_excel_file(results, result_dir)
        
        # 5. 生成折线图
        self._generate_line_chart(results, result_dir)
        
        # 6. 生成累计收益折线图
        self._generate_cumulative_returns_chart(results, result_dir)
        
        print("=" * 80)
    
    def _save_markdown_report(self, results: List[Dict[str, Any]], result_dir: str = None):
        """保存Markdown格式的报告到文件"""
        if not results:
            print("❌ 无数据可保存")
            return
        
        # 获取行业名称
        industry_name = results[0].get('industry_name', 'Unknown')
        
        # 如果没有提供目录，创建新的回测结果目录
        if result_dir is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            category = self.get_industry_category(industry_name)
            result_dir = f"tests/backtest/{category}/{industry_name}/{timestamp}"
            os.makedirs(result_dir, exist_ok=True)
        
        # 生成文件名
        filename = f"{result_dir}/回测报告.md"
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                # 写入报告标题
                f.write(f"# 策略回测结果报告\n\n")
                f.write(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"**行业板块**: {industry_name}\n")
                f.write(f"**回测期间**: {results[0]['start_date']} 至 {results[0]['end_date']}\n")
                f.write(f"**策略数量**: {len(results)}\n\n")
                
                # 写入综合结果表
                f.write("## 📊 综合结果表\n\n")
                comprehensive_data = self._get_comprehensive_data(results)
                if comprehensive_data:
                    comprehensive_df = pd.DataFrame(comprehensive_data)
                    f.write(comprehensive_df.to_markdown(index=False))
                    f.write("\n\n")
                
                # 写入日收益明细表
                f.write("## 📅 日收益明细表\n\n")
                daily_data = self._get_daily_returns_data(results)
                if daily_data:
                    daily_df = pd.DataFrame(daily_data)
                    f.write(daily_df.to_markdown(index=False))
                    f.write("\n\n")
                
                # 写入日收益统计摘要
                f.write("## 📊 日收益统计摘要\n\n")
                summary_data = self._get_daily_returns_summary_data(daily_data, results)
                if summary_data:
                    summary_df = pd.DataFrame(summary_data)
                    f.write(summary_df.to_markdown(index=False))
                    f.write("\n\n")
                
                # 写入累计收益明细表
                f.write("## 📈 累计收益明细表\n\n")
                cumulative_data = self._get_cumulative_returns_data(results)
                if cumulative_data:
                    cumulative_df = pd.DataFrame(cumulative_data)
                    f.write(cumulative_df.to_markdown(index=False))
                    f.write("\n\n")
                
                # 写入累计收益统计摘要
                f.write("## 📊 累计收益统计摘要\n\n")
                cumulative_summary_data = self._get_cumulative_returns_summary_data(cumulative_data, results)
                if cumulative_summary_data:
                    cumulative_summary_df = pd.DataFrame(cumulative_summary_data)
                    f.write(cumulative_summary_df.to_markdown(index=False))
                    f.write("\n\n")
                
                # 写入分析结论
                f.write("## 📈 分析结论\n\n")
                f.write(self._generate_analysis_conclusion(results))
                f.write("\n")
            
            print(f"✅ Markdown报告已保存到: {filename}")
            print(f"📁 回测结果目录: {result_dir}")
            
        except Exception as e:
            print(f"❌ 保存Markdown报告失败: {e}")
    
    def _get_comprehensive_data(self, results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """获取综合结果表数据"""
        comprehensive_data = []
        
        # 添加板块实际表现数据
        if results:
            first_result = results[0]
            industry_name = first_result.get('industry_name', '')
            start_date = first_result.get('start_date', '')
            end_date = first_result.get('end_date', '')
            
            # 获取板块原始历史数据
            hist_data = self.get_historical_data(industry_name, start_date, end_date)
            if hist_data is not None and not hist_data.empty:
                # 找到收盘价列
                close_col = None
                for col in ['收盘价', 'close', 'Close']:
                    if col in hist_data.columns:
                        close_col = col
                        break
                
                if close_col is not None:
                    # 计算板块买入并持有的表现
                    initial_price = hist_data.iloc[0][close_col]
                    final_price = hist_data.iloc[-1][close_col]
                    sector_total_return = (final_price - initial_price) / initial_price
                    
                    # 计算板块年化收益率
                    data_points = len(hist_data)
                    sector_annualized_return = (1 + sector_total_return) ** (252 / data_points) - 1
                    
                    # 计算板块波动率
                    returns = hist_data[close_col].pct_change().dropna()
                    sector_volatility = returns.std() * np.sqrt(252)
                    
                    # 计算板块夏普比率
                    sector_sharpe_ratio = sector_annualized_return / sector_volatility if sector_volatility > 0 else 0
                    
                    # 计算板块最大回撤
                    sector_max_drawdown = self.calculate_max_drawdown(hist_data[close_col])
                    
                    # 计算初始和最终价值（假设初始资金100000）
                    initial_capital = 100000
                    final_value = initial_capital * (1 + sector_total_return)
                    
                    # 添加板块实际表现行
                    comprehensive_data.append({
                        '策略名称': '板块实际表现',
                        '初始资金': f"¥{initial_capital:,.0f}",
                        '最终价值': f"¥{final_value:,.0f}",
                        '总收益率': f"{sector_total_return:.2%}",
                        '年化收益率': f"{sector_annualized_return:.2%}",
                        '波动率': f"{sector_volatility:.2%}",
                        '夏普比率': f"{sector_sharpe_ratio:.4f}",
                        '最大回撤': f"{sector_max_drawdown:.2%}",
                        '总交易次数': 'N/A',
                        '买入次数': 'N/A',
                        '卖出次数': 'N/A',
                        '总交易金额': 'N/A',
                        '平均交易金额': 'N/A',
                        '交易频率': 'N/A',
                        '数据点数': data_points
                    })
        
        # 添加各策略数据
        for result in results:
            trades = result.get('trades', [])
            buy_trades = len([t for t in trades if t['action'] in ['BUY', 'STRONG_BUY']])
            sell_trades = len([t for t in trades if t['action'] in ['SELL', 'STRONG_SELL']])
            
            # 计算平均交易金额
            total_trade_amount = sum(trade['amount'] for trade in trades)
            avg_trade_amount = total_trade_amount / len(trades) if trades else 0
            
            comprehensive_data.append({
                '策略名称': result['strategy_name'],
                '初始资金': f"¥{result['initial_capital']:,.0f}",
                '最终价值': f"¥{result['final_value']:,.0f}",
                '总收益率': f"{result['total_return']:.2%}",
                '年化收益率': f"{result['annualized_return']:.2%}",
                '波动率': f"{result['volatility']:.2%}",
                '夏普比率': f"{result['sharpe_ratio']:.4f}",
                '最大回撤': f"{result['max_drawdown']:.2%}",
                '总交易次数': result['total_trades'],
                '买入次数': buy_trades,
                '卖出次数': sell_trades,
                '总交易金额': f"¥{total_trade_amount:,.0f}",
                '平均交易金额': f"¥{avg_trade_amount:,.0f}",
                '交易频率': f"{result['total_trades'] / result['data_points']:.2f}" if result['data_points'] > 0 else "0.00",
                '数据点数': result['data_points']
            })
        
        return comprehensive_data
    
    def _get_daily_returns_data(self, results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """获取日收益明细数据"""
        if not results:
            return []
        
        # 获取第一个结果的数据作为基准
        first_result = results[0]
        industry_name = first_result.get('industry_name', '')
        start_date = first_result.get('start_date', '')
        end_date = first_result.get('end_date', '')
        
        # 获取板块原始历史数据
        hist_data = self.get_historical_data(industry_name, start_date, end_date)
        if hist_data is None or hist_data.empty:
            return []
        
        # 找到收盘价列
        close_col = None
        for col in ['收盘价', 'close', 'Close']:
            if col in hist_data.columns:
                close_col = col
                break
        
        if close_col is None:
            return []
        
        # 计算板块实际日收益率
        hist_data['板块实际收益率'] = hist_data[close_col].pct_change() * 100
        
        # 准备日收益明细数据
        daily_data = []
        
        for _, row in hist_data.iterrows():
            date_str = row.get('日期', f'Day_{_}')
            if '日期' not in row:
                # 如果没有日期列，使用索引
                date_str = f'Day_{_}'
            
            sector_return = row['板块实际收益率']
            
            # 初始化行数据
            row_data = {
                '日期': date_str,
                '板块实际收益率': f"{sector_return:.2f}%" if not pd.isna(sector_return) else "0.00%"
            }
            
            # 添加每个策略的日收益率
            for result in results:
                strategy_name = result['strategy_name']
                strategy_portfolio = result.get('portfolio_values', [])
                
                # 计算策略日收益率
                strategy_daily_return = "0.00%"
                
                if strategy_portfolio and len(strategy_portfolio) > 0:
                    # 通过索引匹配，因为portfolio_values和历史数据的顺序是一致的
                    current_portfolio = None
                    prev_portfolio = None
                    
                    if _ < len(strategy_portfolio):
                        current_portfolio = strategy_portfolio[_]
                        # 找到前一个交易日
                        if _ > 0:
                            prev_portfolio = strategy_portfolio[_ - 1]
                        
                        # 计算日收益率
                        if current_portfolio and prev_portfolio:
                            current_value = current_portfolio['portfolio_value']
                            prev_value = prev_portfolio['portfolio_value']
                            
                            if prev_value != 0:
                                daily_return = (current_value - prev_value) / prev_value * 100
                                strategy_daily_return = f"{daily_return:.2f}%"
                
                row_data[f'{strategy_name}收益率'] = strategy_daily_return
            
            daily_data.append(row_data)
        
        return daily_data
    
    def _get_cumulative_returns_data(self, results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """获取每日累计收益率数据"""
        if not results:
            return []
        
        # 获取第一个结果的数据作为基准
        first_result = results[0]
        industry_name = first_result.get('industry_name', '')
        start_date = first_result.get('start_date', '')
        end_date = first_result.get('end_date', '')
        
        # 获取板块原始历史数据
        hist_data = self.get_historical_data(industry_name, start_date, end_date)
        if hist_data is None or hist_data.empty:
            return []
        
        # 找到收盘价列
        close_col = None
        for col in ['收盘价', 'close', 'Close']:
            if col in hist_data.columns:
                close_col = col
                break
        
        if close_col is None:
            return []
        
        # 计算板块实际累计收益率
        hist_data['板块实际收益率'] = hist_data[close_col].pct_change() * 100
        hist_data['板块累计收益率'] = (hist_data[close_col] / hist_data[close_col].iloc[0] - 1) * 100
        
        # 准备累计收益明细数据
        cumulative_data = []
        
        for _, row in hist_data.iterrows():
            date_str = row.get('日期', f'Day_{_}')
            if '日期' not in row:
                # 如果没有日期列，使用索引
                date_str = f'Day_{_}'
            
            sector_cumulative = row['板块累计收益率']
            
            # 初始化行数据
            row_data = {
                '日期': date_str,
                '板块累计收益率': f"{sector_cumulative:.2f}%" if not pd.isna(sector_cumulative) else "0.00%"
            }
            
            # 添加每个策略的累计收益率
            for result in results:
                strategy_name = result['strategy_name']
                strategy_portfolio = result.get('portfolio_values', [])
                
                # 计算策略累计收益率
                strategy_cumulative_return = "0.00%"
                
                if strategy_portfolio and len(strategy_portfolio) > 0:
                    # 通过索引匹配，因为portfolio_values和历史数据的顺序是一致的
                    current_portfolio = None
                    
                    if _ < len(strategy_portfolio):
                        current_portfolio = strategy_portfolio[_]
                        
                        # 计算累计收益率
                        if current_portfolio:
                            initial_capital = result.get('initial_capital', 100000)
                            current_value = current_portfolio['portfolio_value']
                            
                            if initial_capital != 0:
                                cumulative_return = (current_value - initial_capital) / initial_capital * 100
                                strategy_cumulative_return = f"{cumulative_return:.2f}%"
                
                row_data[f'{strategy_name}累计收益率'] = strategy_cumulative_return
            
            cumulative_data.append(row_data)
        
        return cumulative_data
    
    def _get_daily_returns_summary_data(self, daily_data: List[Dict[str, Any]], results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """获取日收益统计摘要数据"""
        if not daily_data:
            return []
        
        summary_data = []
        
        # 板块实际收益率统计
        sector_returns = []
        for val in daily_data:
            try:
                sector_returns.append(float(val['板块实际收益率'].replace('%', '')))
            except:
                sector_returns.append(0.0)
        
        summary_data.append({
            '指标': '板块实际收益率',
            '平均日收益率': f"{sum(sector_returns) / len(sector_returns):.2f}%",
            '最大日收益率': f"{max(sector_returns):.2f}%",
            '最小日收益率': f"{min(sector_returns):.2f}%",
            '正收益天数': f"{len([r for r in sector_returns if r > 0])}天",
            '负收益天数': f"{len([r for r in sector_returns if r < 0])}天"
        })
        
        # 各策略收益率统计
        for result in results:
            strategy_name = result['strategy_name']
            col_name = f'{strategy_name}收益率'
            
            if col_name in daily_data[0] if daily_data else False:
                strategy_returns = []
                for val in daily_data:
                    try:
                        strategy_returns.append(float(val[col_name].replace('%', '')))
                    except:
                        strategy_returns.append(0.0)
                
                summary_data.append({
                    '指标': f'{strategy_name}收益率',
                    '平均日收益率': f"{sum(strategy_returns) / len(strategy_returns):.2f}%",
                    '最大日收益率': f"{max(strategy_returns):.2f}%",
                    '最小日收益率': f"{min(strategy_returns):.2f}%",
                    '正收益天数': f"{len([r for r in strategy_returns if r > 0])}天",
                    '负收益天数': f"{len([r for r in strategy_returns if r < 0])}天"
                })
        
        return summary_data
    
    def _generate_analysis_conclusion(self, results: List[Dict[str, Any]]) -> str:
        """生成分析结论"""
        if not results:
            return "无数据可分析"
        
        conclusion = []
        
        # 找出表现最好的策略
        best_strategy = max(results, key=lambda x: x['total_return'])
        worst_strategy = min(results, key=lambda x: x['total_return'])
        
        conclusion.append(f"### 策略表现分析\n")
        conclusion.append(f"- **最佳策略**: {best_strategy['strategy_name']} (总收益率: {best_strategy['total_return']:.2%})\n")
        conclusion.append(f"- **最差策略**: {worst_strategy['strategy_name']} (总收益率: {worst_strategy['total_return']:.2%})\n")
        
        # 交易活跃度分析
        active_strategies = [r for r in results if r['total_trades'] > 0]
        inactive_strategies = [r for r in results if r['total_trades'] == 0]
        
        conclusion.append(f"### 交易活跃度分析\n")
        conclusion.append(f"- **活跃策略**: {len(active_strategies)} 个\n")
        conclusion.append(f"- **非活跃策略**: {len(inactive_strategies)} 个\n")
        
        if active_strategies:
            most_active = max(active_strategies, key=lambda x: x['total_trades'])
            conclusion.append(f"- **最活跃策略**: {most_active['strategy_name']} (交易次数: {most_active['total_trades']})\n")
        
        # 风险分析
        conclusion.append(f"### 风险分析\n")
        for result in results:
            conclusion.append(f"- **{result['strategy_name']}**: 最大回撤 {result['max_drawdown']:.2%}, 夏普比率 {result['sharpe_ratio']:.4f}\n")
        
        return "".join(conclusion)
    
    def _generate_excel_file(self, results: List[Dict[str, Any]], result_dir: str = None):
        """生成Excel格式的日收益明细表文件"""
        if not results:
            print("❌ 无数据可生成Excel文件")
            return
        
        try:
            # 导入openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # 获取行业名称
            industry_name = results[0].get('industry_name', 'Unknown')
            
            # 如果没有提供目录，创建新的回测结果目录
            if result_dir is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                result_dir = f"backtest/{industry_name}_{timestamp}"
                os.makedirs(result_dir, exist_ok=True)
            
            # 生成文件名
            filename = f"{result_dir}/收益率数据明细表.xlsx"
            
            # 获取日收益数据
            daily_data = self._get_daily_returns_data(results)
            if not daily_data:
                print("❌ 无日收益数据可生成Excel文件")
                return
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "日收益明细表"
            
            # 准备数据
            headers = ['日期', '板块实际收益率']
            
            # 添加各策略收益率列
            for result in results:
                strategy_name = result['strategy_name']
                headers.append(f'{strategy_name}收益率')
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # 写入数据
            for row_idx, row_data in enumerate(daily_data, 2):
                # 写入日期
                ws.cell(row=row_idx, column=1, value=row_data['日期'])
                
                # 写入板块实际收益率
                sector_return = row_data['板块实际收益率']
                ws.cell(row=row_idx, column=2, value=sector_return)
                
                # 写入各策略收益率
                for col_idx, result in enumerate(results, 3):
                    strategy_name = result['strategy_name']
                    strategy_return = row_data[f'{strategy_name}收益率']
                    ws.cell(row=row_idx, column=col_idx, value=strategy_return)
            
            # 设置列宽
            ws.column_dimensions['A'].width = 15  # 日期列
            for col in range(2, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 18  # 收益率列
            
            # 添加数据验证和格式化
            for row in range(2, len(daily_data) + 2):
                for col in range(2, len(headers) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 根据正负值设置颜色
                    try:
                        value = float(cell.value.replace('%', ''))
                        if value > 0:
                            cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # 浅绿色
                        elif value < 0:
                            cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # 浅红色
                    except:
                        pass
            
            # 创建累计收益明细工作表
            cumulative_data = self._get_cumulative_returns_data(results)
            if cumulative_data:
                cumulative_ws = wb.create_sheet("累计收益明细表")
                
                # 准备累计收益数据
                cumulative_headers = ['日期', '板块累计收益率']
                
                # 添加各策略累计收益率列
                for result in results:
                    strategy_name = result['strategy_name']
                    cumulative_headers.append(f'{strategy_name}累计收益率')
                
                # 写入累计收益表头
                for col, header in enumerate(cumulative_headers, 1):
                    cell = cumulative_ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, size=12)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                
                # 写入累计收益数据
                for row_idx, row_data in enumerate(cumulative_data, 2):
                    # 写入日期
                    cumulative_ws.cell(row=row_idx, column=1, value=row_data['日期'])
                    
                    # 写入板块累计收益率
                    sector_cumulative = row_data['板块累计收益率']
                    cumulative_ws.cell(row=row_idx, column=2, value=sector_cumulative)
                    
                    # 写入各策略累计收益率
                    for col_idx, result in enumerate(results, 3):
                        strategy_name = result['strategy_name']
                        strategy_cumulative = row_data[f'{strategy_name}累计收益率']
                        cumulative_ws.cell(row=row_idx, column=col_idx, value=strategy_cumulative)
                
                # 设置累计收益列宽
                cumulative_ws.column_dimensions['A'].width = 15  # 日期列
                for col in range(2, len(cumulative_headers) + 1):
                    cumulative_ws.column_dimensions[chr(64 + col)].width = 18  # 收益率列
                
                # 添加累计收益数据验证和格式化
                for row in range(2, len(cumulative_data) + 2):
                    for col in range(2, len(cumulative_headers) + 1):
                        cell = cumulative_ws.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # 根据正负值设置颜色
                        try:
                            value = float(cell.value.replace('%', ''))
                            if value > 0:
                                cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # 浅绿色
                            elif value < 0:
                                cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # 浅红色
                        except:
                            pass
            
            # 添加汇总信息工作表
            summary_ws = wb.create_sheet("汇总信息")
            
            # 写入基本信息
            summary_ws['A1'] = "回测基本信息"
            summary_ws['A1'].font = Font(bold=True, size=14)
            
            summary_ws['A3'] = "生成时间"
            summary_ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            summary_ws['A4'] = "行业板块"
            summary_ws['B4'] = industry_name
            summary_ws['A5'] = "回测期间"
            summary_ws['B5'] = f"{results[0]['start_date']} 至 {results[0]['end_date']}"
            summary_ws['A6'] = "策略数量"
            summary_ws['B6'] = len(results)
            
            # 写入策略汇总
            summary_ws['A8'] = "策略汇总"
            summary_ws['A8'].font = Font(bold=True, size=12)
            
            summary_headers = ['策略名称', '总收益率', '年化收益率', '夏普比率', '最大回撤', '总交易次数']
            for col, header in enumerate(summary_headers, 1):
                cell = summary_ws.cell(row=9, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # 写入策略数据
            for row_idx, result in enumerate(results, 10):
                summary_ws.cell(row=row_idx, column=1, value=result['strategy_name'])
                summary_ws.cell(row=row_idx, column=2, value=f"{result['total_return']:.2%}")
                summary_ws.cell(row=row_idx, column=3, value=f"{result['annualized_return']:.2%}")
                summary_ws.cell(row=row_idx, column=4, value=f"{result['sharpe_ratio']:.4f}")
                summary_ws.cell(row=row_idx, column=5, value=f"{result['max_drawdown']:.2%}")
                summary_ws.cell(row=row_idx, column=6, value=result['total_trades'])
            
            # 设置汇总工作表列宽
            for col in range(1, len(summary_headers) + 1):
                summary_ws.column_dimensions[chr(64 + col)].width = 15
            
            # 保存文件
            wb.save(filename)
            print(f"✅ Excel文件已保存到: {filename}")
            print(f"📁 回测结果目录: {result_dir}")
            
        except ImportError:
            print("❌ 需要安装openpyxl库: pip install openpyxl")
        except Exception as e:
            print(f"❌ 生成Excel文件失败: {e}")
    
    def _generate_line_chart(self, results: List[Dict[str, Any]], result_dir: str = None):
        """生成日收益明细表的折线图"""
        if not results:
            print("❌ 无数据可生成折线图")
            return
        
        try:
            # 导入matplotlib
            import matplotlib.pyplot as plt
            import matplotlib.dates as mdates
            from datetime import datetime
            import warnings
            
            # 设置中文字体支持
            plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS', 'DejaVu Sans', 'Arial']
            plt.rcParams['axes.unicode_minus'] = False
            
            # 忽略字体警告
            warnings.filterwarnings('ignore', category=UserWarning, module='matplotlib')
            
            # 获取行业名称
            industry_name = results[0].get('industry_name', 'Unknown')
            
            # 如果没有提供目录，创建新的回测结果目录
            if result_dir is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                result_dir = f"backtest/{industry_name}_{timestamp}"
                os.makedirs(result_dir, exist_ok=True)
            
            # 生成文件名
            filename = f"{result_dir}/每日收益率折线图.png"
            
            # 获取日收益数据
            daily_data = self._get_daily_returns_data(results)
            if not daily_data:
                print("❌ 无日收益数据可生成折线图")
                return
            
            # 数据读取与预处理
            df = pd.DataFrame(daily_data)
            
            # 处理日期列 - 转换为datetime格式
            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')
            
            # 处理百分比数据 - 去除%符号并转换为浮点数
            numeric_columns = ['板块实际收益率']
            for result in results:
                strategy_name = result['strategy_name']
                numeric_columns.append(f'{strategy_name}收益率')
            
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = df[col].str.replace('%', '').astype(float)
            
            # 可视化处理
            plt.figure(figsize=(14, 8))  # 设置合适的图表尺寸
            
            # 绘制多系列折线图
            colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b']
            
            # 绘制板块实际收益率
            plt.plot(df['日期'], df['板块实际收益率'], 
                    label='板块实际收益率', linewidth=2.5, color=colors[0], marker='o', markersize=6)
            
            # 绘制各策略收益率
            for i, result in enumerate(results):
                strategy_name = result['strategy_name']
                col_name = f'{strategy_name}收益率'
                if col_name in df.columns:
                    plt.plot(df['日期'], df[col_name], 
                            label=f'{strategy_name}收益率', linewidth=2, 
                            color=colors[(i + 1) % len(colors)], marker='s', markersize=5)
            
            # 添加图表元素
            plt.title(f'{industry_name}板块日收益率对比图', fontsize=16, fontweight='bold', pad=20)
            plt.xlabel('日期', fontsize=12, fontweight='bold')
            plt.ylabel('收益率 (%)', fontsize=12, fontweight='bold')
            
            # 添加图例
            plt.legend(loc='upper left', frameon=True, fancybox=True, shadow=True)
            
            # 调整x轴标签旋转角度(45度)避免重叠
            plt.xticks(rotation=45)
            
            # 添加网格线提高可读性
            plt.grid(True, alpha=0.3, linestyle='--')
            
            # 设置y轴零线
            plt.axhline(y=0, color='black', linestyle='-', alpha=0.5, linewidth=1)
            
            # 调整布局
            plt.tight_layout()
            
            # 保存图表
            plt.savefig(filename, dpi=300, bbox_inches='tight', facecolor='white')
            plt.close()  # 关闭图表以释放内存
            
            print(f"✅ 折线图已保存到: {filename}")
            print(f"📁 回测结果目录: {result_dir}")
            
        except ImportError:
            print("❌ 需要安装matplotlib库: pip install matplotlib")
        except Exception as e:
            print(f"❌ 生成折线图失败: {e}")
    
    def _generate_cumulative_returns_chart(self, results: List[Dict[str, Any]], result_dir: str = None):
        """生成累计收益明细表的折线图"""
        if not results:
            print("❌ 无数据可生成累计收益折线图")
            return
        
        try:
            # 导入matplotlib
            import matplotlib.pyplot as plt
            import matplotlib.dates as mdates
            from datetime import datetime
            import warnings
            
            # 设置中文字体支持
            plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS', 'DejaVu Sans', 'Arial']
            plt.rcParams['axes.unicode_minus'] = False
            
            # 忽略字体警告
            warnings.filterwarnings('ignore', category=UserWarning, module='matplotlib')
            
            # 获取行业名称
            industry_name = results[0].get('industry_name', 'Unknown')
            
            # 如果没有提供目录，创建新的回测结果目录
            if result_dir is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                result_dir = f"backtest/{industry_name}_{timestamp}"
                os.makedirs(result_dir, exist_ok=True)
            
            # 生成文件名
            filename = f"{result_dir}/累计收益率折线图.png"
            
            # 获取累计收益数据
            cumulative_data = self._get_cumulative_returns_data(results)
            if not cumulative_data:
                print("❌ 无累计收益数据可生成折线图")
                return
            
            # 数据读取与预处理
            df = pd.DataFrame(cumulative_data)
            
            # 处理日期列 - 转换为datetime格式
            df['日期'] = pd.to_datetime(df['日期'], errors='coerce')
            
            # 处理百分比数据 - 去除%符号并转换为浮点数
            numeric_columns = ['板块累计收益率']
            for result in results:
                strategy_name = result['strategy_name']
                numeric_columns.append(f'{strategy_name}累计收益率')
            
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = df[col].str.replace('%', '').astype(float)
            
            # 可视化处理
            plt.figure(figsize=(14, 8))  # 设置合适的图表尺寸
            
            # 绘制多系列折线图
            colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b']
            
            # 绘制板块累计收益率
            plt.plot(df['日期'], df['板块累计收益率'], 
                    label='板块累计收益率', linewidth=2.5, color=colors[0], marker='o', markersize=6)
            
            # 绘制各策略累计收益率
            for i, result in enumerate(results):
                strategy_name = result['strategy_name']
                col_name = f'{strategy_name}累计收益率'
                if col_name in df.columns:
                    plt.plot(df['日期'], df[col_name], 
                            label=f'{strategy_name}累计收益率', linewidth=2, 
                            color=colors[(i + 1) % len(colors)], marker='s', markersize=5)
            
            # 添加图表元素
            plt.title(f'{industry_name}板块累计收益率对比图', fontsize=16, fontweight='bold', pad=20)
            plt.xlabel('日期', fontsize=12, fontweight='bold')
            plt.ylabel('累计收益率 (%)', fontsize=12, fontweight='bold')
            
            # 添加图例
            plt.legend(loc='upper left', frameon=True, fancybox=True, shadow=True)
            
            # 调整x轴标签旋转角度(45度)避免重叠
            plt.xticks(rotation=45)
            
            # 添加网格线提高可读性
            plt.grid(True, alpha=0.3, linestyle='--')
            
            # 设置y轴零线
            plt.axhline(y=0, color='black', linestyle='-', alpha=0.5, linewidth=1)
            
            # 调整布局
            plt.tight_layout()
            
            # 保存图表
            plt.savefig(filename, dpi=300, bbox_inches='tight', facecolor='white')
            plt.close()  # 关闭图表以释放内存
            
            print(f"✅ 累计收益折线图已保存到: {filename}")
            print(f"📁 回测结果目录: {result_dir}")
            
        except ImportError:
            print("❌ 需要安装matplotlib库: pip install matplotlib")
        except Exception as e:
            print(f"❌ 生成累计收益折线图失败: {e}")
    
    def _print_cumulative_returns_details(self, results: List[Dict[str, Any]]):
        """打印累计收益明细表"""
        print("📈 累计收益明细表")
        print("-" * 80)
        
        if not results:
            print("无数据可显示")
            return
        
        cumulative_data = self._get_cumulative_returns_data(results)
        if not cumulative_data:
            print("无累计收益数据可显示")
            return
        
        cumulative_df = pd.DataFrame(cumulative_data)
        print(cumulative_df.to_string(index=False))
        
        print("\n📊 累计收益统计摘要")
        print("-" * 80)
        
        summary_data = self._get_cumulative_returns_summary_data(cumulative_data, results)
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            print(summary_df.to_string(index=False))
        print()
    
    def _print_comprehensive_table(self, results: List[Dict[str, Any]]):
        """打印综合结果表（合并汇总、详细、交易统计）"""
        print("📊 综合结果表")
        print("-" * 120)
        
        # 准备综合数据
        comprehensive_data = []
        
        # 添加板块实际表现数据
        if results:
            first_result = results[0]
            industry_name = first_result.get('industry_name', '')
            start_date = first_result.get('start_date', '')
            end_date = first_result.get('end_date', '')
            
            # 获取板块原始历史数据
            hist_data = self.get_historical_data(industry_name, start_date, end_date)
            if hist_data is not None and not hist_data.empty:
                # 找到收盘价列
                close_col = None
                for col in ['收盘价', 'close', 'Close']:
                    if col in hist_data.columns:
                        close_col = col
                        break
                
                if close_col is not None:
                    # 计算板块买入并持有的表现
                    initial_price = hist_data.iloc[0][close_col]
                    final_price = hist_data.iloc[-1][close_col]
                    sector_total_return = (final_price - initial_price) / initial_price
                    
                    # 计算板块年化收益率
                    data_points = len(hist_data)
                    sector_annualized_return = (1 + sector_total_return) ** (252 / data_points) - 1
                    
                    # 计算板块波动率
                    returns = hist_data[close_col].pct_change().dropna()
                    sector_volatility = returns.std() * np.sqrt(252)
                    
                    # 计算板块夏普比率
                    sector_sharpe_ratio = sector_annualized_return / sector_volatility if sector_volatility > 0 else 0
                    
                    # 计算板块最大回撤
                    sector_max_drawdown = self.calculate_max_drawdown(hist_data[close_col])
                    
                    # 计算初始和最终价值（假设初始资金100000）
                    initial_capital = 100000
                    final_value = initial_capital * (1 + sector_total_return)
                    
                    # 添加板块实际表现行
                    comprehensive_data.append({
                        '策略名称': '板块实际表现',
                        '初始资金': f"¥{initial_capital:,.0f}",
                        '最终价值': f"¥{final_value:,.0f}",
                        '总收益率': f"{sector_total_return:.2%}",
                        '年化收益率': f"{sector_annualized_return:.2%}",
                        '波动率': f"{sector_volatility:.2%}",
                        '夏普比率': f"{sector_sharpe_ratio:.4f}",
                        '最大回撤': f"{sector_max_drawdown:.2%}",
                        '总交易次数': 'N/A',
                        '买入次数': 'N/A',
                        '卖出次数': 'N/A',
                        '总交易金额': 'N/A',
                        '平均交易金额': 'N/A',
                        '交易频率': 'N/A',
                        '数据点数': data_points
                    })
        
        # 添加各策略数据
        for result in results:
            trades = result.get('trades', [])
            buy_trades = len([t for t in trades if t['action'] in ['BUY', 'STRONG_BUY']])
            sell_trades = len([t for t in trades if t['action'] in ['SELL', 'STRONG_SELL']])
            
            # 计算平均交易金额
            total_trade_amount = sum(trade['amount'] for trade in trades)
            avg_trade_amount = total_trade_amount / len(trades) if trades else 0
            
            comprehensive_data.append({
                '策略名称': result['strategy_name'],
                '初始资金': f"¥{result['initial_capital']:,.0f}",
                '最终价值': f"¥{result['final_value']:,.0f}",
                '总收益率': f"{result['total_return']:.2%}",
                '年化收益率': f"{result['annualized_return']:.2%}",
                '波动率': f"{result['volatility']:.2%}",
                '夏普比率': f"{result['sharpe_ratio']:.4f}",
                '最大回撤': f"{result['max_drawdown']:.2%}",
                '总交易次数': result['total_trades'],
                '买入次数': buy_trades,
                '卖出次数': sell_trades,
                '总交易金额': f"¥{total_trade_amount:,.0f}",
                '平均交易金额': f"¥{avg_trade_amount:,.0f}",
                '交易频率': f"{result['total_trades'] / result['data_points']:.2f}" if result['data_points'] > 0 else "0.00",
                '数据点数': result['data_points']
            })
        
        comprehensive_df = pd.DataFrame(comprehensive_data)
        
        # 设置pandas显示选项
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.max_colwidth', 12)
        
        print(comprehensive_df.to_string(index=False))
        print()
    
    def _print_summary_table(self, results: List[Dict[str, Any]]):
        """打印策略汇总表"""
        print("📊 策略汇总表")
        print("-" * 80)
        
        # 准备汇总数据
        summary_data = []
        
        # 添加板块实际表现数据
        if results:
            first_result = results[0]
            industry_name = first_result.get('industry_name', '')
            start_date = first_result.get('start_date', '')
            end_date = first_result.get('end_date', '')
            
            # 获取板块原始历史数据
            hist_data = self.get_historical_data(industry_name, start_date, end_date)
            if hist_data is not None and not hist_data.empty:
                # 找到收盘价列
                close_col = None
                for col in ['收盘价', 'close', 'Close']:
                    if col in hist_data.columns:
                        close_col = col
                        break
                
                if close_col is not None:
                    # 计算板块买入并持有的表现
                    initial_price = hist_data.iloc[0][close_col]
                    final_price = hist_data.iloc[-1][close_col]
                    sector_total_return = (final_price - initial_price) / initial_price
                    
                    # 计算板块年化收益率
                    data_points = len(hist_data)
                    sector_annualized_return = (1 + sector_total_return) ** (252 / data_points) - 1
                    
                    # 计算板块波动率
                    returns = hist_data[close_col].pct_change().dropna()
                    sector_volatility = returns.std() * np.sqrt(252)
                    
                    # 计算板块夏普比率
                    sector_sharpe_ratio = sector_annualized_return / sector_volatility if sector_volatility > 0 else 0
                    
                    # 计算板块最大回撤
                    sector_max_drawdown = self.calculate_max_drawdown(hist_data[close_col])
                    
                    # 计算最终价值（假设初始资金100000）
                    initial_capital = 100000
                    final_value = initial_capital * (1 + sector_total_return)
                    
                    # 添加板块实际表现行
                    summary_data.append({
                        '策略名称': '板块实际表现',
                        '总收益率': f"{sector_total_return:.2%}",
                        '年化收益率': f"{sector_annualized_return:.2%}",
                        '夏普比率': f"{sector_sharpe_ratio:.4f}",
                        '最大回撤': f"{sector_max_drawdown:.2%}",
                        '总交易次数': 'N/A',
                        '最终价值': f"¥{final_value:,.0f}"
                    })
        
        # 添加各策略数据
        for result in results:
            summary_data.append({
                '策略名称': result['strategy_name'],
                '总收益率': f"{result['total_return']:.2%}",
                '年化收益率': f"{result['annualized_return']:.2%}",
                '夏普比率': f"{result['sharpe_ratio']:.4f}",
                '最大回撤': f"{result['max_drawdown']:.2%}",
                '总交易次数': result['total_trades'],
                '最终价值': f"¥{result['final_value']:,.0f}"
            })
        
        summary_df = pd.DataFrame(summary_data)
        
        # 设置pandas显示选项
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.max_colwidth', 20)
        
        print(summary_df.to_string(index=False))
        print()
    
    def _print_detailed_results(self, results: List[Dict[str, Any]]):
        """打印详细结果表"""
        print("📈 详细结果表")
        print("-" * 80)
        
        # 准备详细数据
        detailed_data = []
        
        # 添加板块实际表现数据
        if results:
            first_result = results[0]
            industry_name = first_result.get('industry_name', '')
            start_date = first_result.get('start_date', '')
            end_date = first_result.get('end_date', '')
            
            # 获取板块原始历史数据
            hist_data = self.get_historical_data(industry_name, start_date, end_date)
            if hist_data is not None and not hist_data.empty:
                # 找到收盘价列
                close_col = None
                for col in ['收盘价', 'close', 'Close']:
                    if col in hist_data.columns:
                        close_col = col
                        break
                
                if close_col is not None:
                    # 计算板块买入并持有的表现
                    initial_price = hist_data.iloc[0][close_col]
                    final_price = hist_data.iloc[-1][close_col]
                    sector_total_return = (final_price - initial_price) / initial_price
                    
                    # 计算板块年化收益率
                    data_points = len(hist_data)
                    sector_annualized_return = (1 + sector_total_return) ** (252 / data_points) - 1
                    
                    # 计算板块波动率
                    returns = hist_data[close_col].pct_change().dropna()
                    sector_volatility = returns.std() * np.sqrt(252)
                    
                    # 计算板块夏普比率
                    sector_sharpe_ratio = sector_annualized_return / sector_volatility if sector_volatility > 0 else 0
                    
                    # 计算板块最大回撤
                    sector_max_drawdown = self.calculate_max_drawdown(hist_data[close_col])
                    
                    # 计算初始和最终价值（假设初始资金100000）
                    initial_capital = 100000
                    final_value = initial_capital * (1 + sector_total_return)
                    
                    # 添加板块实际表现行
                    detailed_data.append({
                        '策略名称': '板块实际表现',
                        '初始资金': f"¥{initial_capital:,.0f}",
                        '最终价值': f"¥{final_value:,.0f}",
                        '总收益率': f"{sector_total_return:.2%}",
                        '年化收益率': f"{sector_annualized_return:.2%}",
                        '波动率': f"{sector_volatility:.2%}",
                        '夏普比率': f"{sector_sharpe_ratio:.4f}",
                        '最大回撤': f"{sector_max_drawdown:.2%}",
                        '数据点数': data_points
                    })
        
        # 添加各策略数据
        for result in results:
            detailed_data.append({
                '策略名称': result['strategy_name'],
                '初始资金': f"¥{result['initial_capital']:,.0f}",
                '最终价值': f"¥{result['final_value']:,.0f}",
                '总收益率': f"{result['total_return']:.2%}",
                '年化收益率': f"{result['annualized_return']:.2%}",
                '波动率': f"{result['volatility']:.2%}",
                '夏普比率': f"{result['sharpe_ratio']:.4f}",
                '最大回撤': f"{result['max_drawdown']:.2%}",
                '数据点数': result['data_points']
            })
        
        detailed_df = pd.DataFrame(detailed_data)
        print(detailed_df.to_string(index=False))
        print()
    
    def _print_trading_statistics(self, results: List[Dict[str, Any]]):
        """打印交易统计表"""
        print("💰 交易统计表")
        print("-" * 80)
        
        # 准备交易统计数据
        trading_data = []
        for result in results:
            trades = result.get('trades', [])
            buy_trades = len([t for t in trades if t['action'] in ['BUY', 'STRONG_BUY']])
            sell_trades = len([t for t in trades if t['action'] in ['SELL', 'STRONG_SELL']])
            
            # 计算平均交易金额
            total_trade_amount = sum(trade['amount'] for trade in trades)
            avg_trade_amount = total_trade_amount / len(trades) if trades else 0
            
            trading_data.append({
                '策略名称': result['strategy_name'],
                '总交易次数': result['total_trades'],
                '买入次数': buy_trades,
                '卖出次数': sell_trades,
                '总交易金额': f"¥{total_trade_amount:,.0f}",
                '平均交易金额': f"¥{avg_trade_amount:,.0f}",
                '交易频率': f"{result['total_trades'] / result['data_points']:.2f}" if result['data_points'] > 0 else "0.00"
            })
        
        trading_df = pd.DataFrame(trading_data)
        print(trading_df.to_string(index=False))
        print()
    
    def _get_cumulative_returns_summary_data(self, cumulative_data: List[Dict[str, Any]], results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """获取累计收益统计摘要数据"""
        if not cumulative_data:
            return []
        
        summary_data = []
        
        # 板块实际累计收益率统计
        sector_cumulative = []
        for val in cumulative_data:
            try:
                sector_cumulative.append(float(val['板块累计收益率'].replace('%', '')))
            except:
                sector_cumulative.append(0.0)
        
        summary_data.append({
            '指标': '板块累计收益率',
            '最终累计收益率': f"{sector_cumulative[-1]:.2f}%",
            '最大累计收益率': f"{max(sector_cumulative):.2f}%",
            '最小累计收益率': f"{min(sector_cumulative):.2f}%",
            '累计收益波动': f"{max(sector_cumulative) - min(sector_cumulative):.2f}%",
            '收益稳定性': f"{'稳定' if max(sector_cumulative) - min(sector_cumulative) < 20 else '波动'}"
        })
        
        # 各策略累计收益率统计
        for result in results:
            strategy_name = result['strategy_name']
            col_name = f'{strategy_name}累计收益率'
            
            if col_name in cumulative_data[0] if cumulative_data else False:
                strategy_cumulative = []
                for val in cumulative_data:
                    try:
                        strategy_cumulative.append(float(val[col_name].replace('%', '')))
                    except:
                        strategy_cumulative.append(0.0)
                
                summary_data.append({
                    '指标': f'{strategy_name}累计收益率',
                    '最终累计收益率': f"{strategy_cumulative[-1]:.2f}%",
                    '最大累计收益率': f"{max(strategy_cumulative):.2f}%",
                    '最小累计收益率': f"{min(strategy_cumulative):.2f}%",
                    '累计收益波动': f"{max(strategy_cumulative) - min(strategy_cumulative):.2f}%",
                    '收益稳定性': f"{'稳定' if max(strategy_cumulative) - min(strategy_cumulative) < 20 else '波动'}"
                })
        
        return summary_data
    
    def _print_daily_returns_details(self, results: List[Dict[str, Any]]):
        """打印日收益明细表"""
        print("📅 日收益明细表")
        print("-" * 80)
        
        if not results:
            print("无数据可显示")
            return
        
        # 获取第一个结果的数据作为基准
        first_result = results[0]
        industry_name = first_result.get('industry_name', '')
        start_date = first_result.get('start_date', '')
        end_date = first_result.get('end_date', '')
        
        # 获取板块原始历史数据
        hist_data = self.get_historical_data(industry_name, start_date, end_date)
        if hist_data is None or hist_data.empty:
            print("无数据可显示")
            return
        
        # 找到收盘价列
        close_col = None
        for col in ['收盘价', 'close', 'Close']:
            if col in hist_data.columns:
                close_col = col
                break
        
        if close_col is None:
            print("未找到收盘价列")
            return
        
        # 计算板块实际日收益率
        hist_data['板块实际收益率'] = hist_data[close_col].pct_change() * 100
        
        # 准备日收益明细数据
        daily_data = []
        
        for _, row in hist_data.iterrows():
            date_str = row.get('日期', f'Day_{_}')
            if '日期' not in row:
                # 如果没有日期列，使用索引
                date_str = f'Day_{_}'
            
            sector_return = row['板块实际收益率']
            
            # 初始化行数据
            row_data = {
                '日期': date_str,
                '板块实际收益率': f"{sector_return:.2f}%" if not pd.isna(sector_return) else "0.00%"
            }
            
            # 添加每个策略的日收益率
            for result in results:
                strategy_name = result['strategy_name']
                strategy_portfolio = result.get('portfolio_values', [])
                
                # 计算策略日收益率
                strategy_daily_return = "0.00%"
                
                if strategy_portfolio and len(strategy_portfolio) > 0:
                    # 通过索引匹配，因为portfolio_values和历史数据的顺序是一致的
                    current_portfolio = None
                    prev_portfolio = None
                    
                    if _ < len(strategy_portfolio):
                        current_portfolio = strategy_portfolio[_]
                        # 找到前一个交易日
                        if _ > 0:
                            prev_portfolio = strategy_portfolio[_ - 1]
                        
                        # 计算日收益率
                        if current_portfolio and prev_portfolio:
                            current_value = current_portfolio['portfolio_value']
                            prev_value = prev_portfolio['portfolio_value']
                            
                            if prev_value != 0:
                                daily_return = (current_value - prev_value) / prev_value * 100
                                strategy_daily_return = f"{daily_return:.2f}%"
                
                row_data[f'{strategy_name}收益率'] = strategy_daily_return
            
            daily_data.append(row_data)
        
        # 创建DataFrame并显示
        daily_df = pd.DataFrame(daily_data)
        
        # 设置pandas显示选项
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.max_colwidth', 15)
        
        print(daily_df.to_string(index=False))
        print()
        
        # 添加日收益统计摘要
        self._print_daily_returns_summary(daily_df, results)
    
    def _print_daily_returns_summary(self, daily_df: pd.DataFrame, results: List[Dict[str, Any]]):
        """打印日收益统计摘要"""
        print("📊 日收益统计摘要")
        print("-" * 80)
        
        # 准备统计摘要数据
        summary_data = []
        
        # 板块实际收益率统计
        sector_returns = []
        for val in daily_df['板块实际收益率']:
            try:
                sector_returns.append(float(val.replace('%', '')))
            except:
                sector_returns.append(0.0)
        
        summary_data.append({
            '指标': '板块实际收益率',
            '平均日收益率': f"{sum(sector_returns) / len(sector_returns):.2f}%",
            '最大日收益率': f"{max(sector_returns):.2f}%",
            '最小日收益率': f"{min(sector_returns):.2f}%",
            '正收益天数': f"{len([r for r in sector_returns if r > 0])}天",
            '负收益天数': f"{len([r for r in sector_returns if r < 0])}天"
        })
        
        # 各策略收益率统计
        for result in results:
            strategy_name = result['strategy_name']
            col_name = f'{strategy_name}收益率'
            
            if col_name in daily_df.columns:
                strategy_returns = []
                for val in daily_df[col_name]:
                    try:
                        strategy_returns.append(float(val.replace('%', '')))
                    except:
                        strategy_returns.append(0.0)
                
                summary_data.append({
                    '指标': f'{strategy_name}收益率',
                    '平均日收益率': f"{sum(strategy_returns) / len(strategy_returns):.2f}%",
                    '最大日收益率': f"{max(strategy_returns):.2f}%",
                    '最小日收益率': f"{min(strategy_returns):.2f}%",
                    '正收益天数': f"{len([r for r in strategy_returns if r > 0])}天",
                    '负收益天数': f"{len([r for r in strategy_returns if r < 0])}天"
                })
        
        summary_df = pd.DataFrame(summary_data)
        print(summary_df.to_string(index=False))
        print()
    
